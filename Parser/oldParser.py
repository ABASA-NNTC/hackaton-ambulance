from openpyxl import load_workbook
import pyexcel
import os
import datetime
import threading
import zipfile


# Функция конвертирует файлы формата .xls в .xlsx
# Файл формата .xlsx необходим для работы openpyxl
def convert_files_to_xlsx():
    if (not os.path.exists("data_xlsx")):
        os.mkdir("data_xlsx")
    for directory in os.listdir(source):
        if not os.path.exists(f'data_xlsx/{directory}'):
            os.mkdir(f'data_xlsx/{directory}')
            for file in os.listdir(f'{source}/{directory}'):
                if (not os.path.exists(f'data_xlsx/{directory}/{file}x')):
                    pyexcel.save_book_as(file_name=f'{source}/{directory}/{file}',
                                         dest_file_name=f'data_xlsx/{directory}/{file}x')
                    print(f'File {file} has been successfully created in {directory}/{file}x!')


def print_data(index, ws):
    print('Индекс:', index)
    print('Дата: ', ws[index][0].value)
    print('Номер: ', ws[index][3].value)
    print('Больной: ', ws[index][5].value)
    print('Возраст: ', ws[index][16].value)
    print('Исполнитель: ', ws[index][19].value)
    
    print('Адрес: ', ws[index + 1][1].value)
    
    print('Повод: ', ws[index + 2][1].value)
    print('Вызов: ', ws[index + 2][11].value)
    print('Вид: ', ws[index + 2][18].value)
    
    print('Диагноз: ', ws[index + 3][1].value)
    print('Результат: ', ws[index + 3][11].value)
    
    print('Доставлен: ', ws[index + 4][1].value)
    print('Подстанция: ', ws[index + 4][17].value)
    
    print('Принят: ', ws[index + 5][8].value)
    print('Приезд: ', ws[index + 5][11].value)
    
    print('\n\n')


file_path = 'Datasets/calls_2020/Журнал-активных-вызовов-01-2020.xlsx'


def convert_to_xlsx(source_path, dest_path):
    pyexcel.save_book_as(file_name=source_path, dest_file_name=dest_path)

def add_columns(worksheet):
    titles = {"A1": "date",
              "B1": "number",
              "C1": "sick",
              "D1": "age",
              "E1": "executor",
              "F1": "address",
              "G1": "occasion",
              "H1": "call",
              "I1": "type",
              "J1": "diagnosis",
              "K1": "result",
              "L1": "delivered",
              "M1": "substation",
              "N1": "accepted",
              "O1": "arrival"}

    for key, value in titles.items():
        worksheet[key] = value


def get_values(worksheet, index):
    cells_with_value = (
        (index, 0),
        (index, 3),
        (index, 5),
        (index, 16),
        (index, 19),
        (index + 1, 1),
        (index + 2, 1),
        (index + 2, 11),
        (index + 2, 18),
        (index + 3, 1),
        (index + 3, 11),
        (index + 4, 1),
        (index + 4, 17),
        (index + 5, 8),
        (index + 5, 11)
    )

    result = []
    for cell in cells_with_value:
        result.append(worksheet[cell[0]][cell[1]].value)
    return result


def get_all_values(source_worksheet):
    source_index = 4
    step_size = 6

    values = []
    while source_index < source_worksheet.max_row:
        if source_worksheet[source_index][0].value:
            if source_worksheet[source_index][2].value:
                source_index += step_size
                values.append(get_values(source_worksheet, source_index))
            else:
                source_index += 1
                while not source_worksheet[source_index][0].value:
                    source_index += 1
        else:
            values.append(get_values(source_worksheet, source_index))
            source_index += step_size

    return values


def create_csv(values, dest_path):
    file = open(dest_path, 'w')
    for row in values:
        file.write(";".join([str(cell) for cell in row]) + "\n")


def parse_file(source_path, dest_path):
    t1 = datetime.datetime.now()
    print(f"{t1}: File '{source_path}' has been started!")
    source_workbook = load_workbook(filename=source_path)
    try:
        source_worksheet = source_workbook.active
        values = get_all_values(source_worksheet)
        create_csv(values, dest_path)
        t2 = datetime.datetime.now()
        print(f"{t2}: File '{source_path}' has been successfully processed with time: {t2 - t1}!")
    finally:
        source_workbook.close()


def parse(source, dest):
    if not os.path.exists(source):
        os.mkdir(source)
    if not os.path.exists(dest):
        os.mkdir(dest)

    for file in os.listdir(source):
        file_path = f"{source}/{file}"
        if not os.path.isfile(file_path):
            continue

        name, extension = os.path.splitext(file_path)
        name = name.split('/')[-1]

        if extension != ".xls":
            continue

        file_xlsx_path = f"{source}/{name}.xlsx"
        if not os.path.exists(file_xlsx_path):
            convert_to_xlsx(file_path, file_xlsx_path)
        else:
            try:
                load_workbook(file_xlsx_path).close()
            except zipfile.BadZipfile:
                convert_to_xlsx(file_path, file_xlsx_path)

        file_parsed_path = f"{dest}/{name}.csv"
        thread = threading.Thread(target=parse_file,
                                  args=(file_xlsx_path, file_parsed_path)
                                  )
        thread.start()


# директория из которой будут считываться файлы
source = "data/input"
# директория в которую будут сохраняться файлы
dest = "data/output"

parse(source, dest)
